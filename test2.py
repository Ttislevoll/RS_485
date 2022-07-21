import tkinter as tk
from tkinter import ANCHOR, BOTTOM, CENTER, END, LEFT, N, NW, RIGHT, SOLID, SUNKEN, SW, TOP, W, ttk, messagebox, filedialog
import ctypes
from turtle import width
import serial
import serial.tools.list_ports
from PIL import ImageTk, Image
import logging
import datetime
import struct
from machine import Machine
from sensor import Sensor
import pickle
from typing import List
import os.path
import time
from new_machine import New_Machine
from new_sensor import New_Sensor
from openpyxl import Workbook, load_workbook

#global variables
ports : list = []
ser = None
machines : List[Machine] = []
labels = {}
treeview_dict = {}

#telegrams for communicating with sensors
broadcast = [0x10,0x7f,0x01,0x09,0x89,0x16]
distance_temp_v03a = [0x10,0x00,0x01,0x4c,0x00,0x16]
temperature_v02a = [0x68,0x09,0x09,0x68,0x00,0x01,0x4c,0x30,0x33,0x5e,0xd1,0x0c,0x04,0x00,0x16]
#read block 0x10
sw_version = [0x68,0x09,0x09,0x68,0x00,0x01,0x4c,0x30,0x33,0x5e,0x10,0x02,0x02,0x00,0x16]
article_number = [0x68,0x09,0x09,0x68,0x00,0x01,0x4c,0x30,0x33,0x5e,0x10,0x04,0x04,0x00,0x16]
serial_number = [0x68,0x09,0x09,0x68,0x00,0x01,0x4c,0x30,0x33,0x5e,0x10,0x10,0x04,0x00,0x16]
description = [0x68,0x09,0x09,0x68,0x00,0x01,0x4c,0x30,0x33,0x5e,0x10,0x28,0x20,0x00,0x16]
#read block 0x20
measuring_unit = [0x68,0x09,0x09,0x68,0x00,0x01,0x4c,0x30,0x33,0x5e,0x20,0x03,0x01,0x00,0x16]
measuring_range = [0x68,0x09,0x09,0x68,0x00,0x01,0x4c,0x30,0x33,0x5e,0x20,0x1c,0x04,0x00,0x16]
measuring_offset = [0x68,0x09,0x09,0x68,0x00,0x01,0x4c,0x30,0x33,0x5e,0x20,0x20,0x04,0x00,0x16]
#assign address
assign_address = [0x68,0x09,0x09,0x68,0x00,0x01,0x43,0x37,0x3e,0x00,0x00,0x00,0x00,0x00,0x16]


#serial read functions

#reads temperature from sensor with version 0.2a or 0.3a
def get_temperature(adr: int) -> float:
    """Reads temperature from sensor

    Args:
        adr (int): Sensor address

    Raises:
        Exception: Checksum is not equal
        Exception: Checksum is not equal

    Returns:
        float: Temperature
    """
    if get_sw_version(adr) == "0.2a":
        temperature_v02a[4] = adr
        temperature_v02a[-2] = get_fcs(temperature_v02a, 4)
        ser.write(bytearray(temperature_v02a))
        received = ser.read(19)
        if get_fcs(received, 4) != received[-2]: raise Exception("Checksum is not equal")
        temp = struct.unpack('i', received[13:17])[0]
        serial_num = str(get_serial_number(adr))
        temp_corr = get_temp_corr()
        a = temp_corr[serial_num][0]
        b = temp_corr[serial_num][1]
        temperature = temp * a + b
    else:
        distance_temp_v03a[1] = adr
        distance_temp_v03a[-2] = get_fcs(distance_temp_v03a, 1)
        ser.write(bytearray(distance_temp_v03a))
        received = ser.read(17)
        if get_fcs(received, 4) != received[-2]: raise Exception("Checksum is not equal")
        temperature = struct.unpack('f', received[11:15])[0]
    temperature = round(temperature, 1)
    return temperature

#reads distance from sensor with version 0.2a or 0.3a
def get_distance(adr : int) -> float:
    """Reads distance from sensor

    Args:
        adr (int): Sensor address

    Raises:
        Exception: Checksum is not equal

    Returns:
        float: Distance
    """
    distance_temp_v03a[1] = adr
    distance_temp_v03a[-2] = get_fcs(distance_temp_v03a, 1)
    bytes = 17
    if get_sw_version(adr) == "0.2a": bytes=13
    ser.write(bytearray(distance_temp_v03a))
    received = ser.read(bytes)
    if get_fcs(received, 4) != received[-2]: raise Exception("checksum is not equal")
    distance = struct.unpack('f', received[7:11])[0]
    distance = round(distance)
    return distance

#returns serial number from sensor
def get_serial_number(adr : int) -> int:
    """Reads serial number from sensor

    Args:
        adr (int): Sensor address

    Raises:
        Exception: Checksum is not equal

    Returns:
        int: Serial number
    """
    serial_number[4] = adr
    serial_number[-2] = get_fcs(serial_number, 4)
    ser.write(bytearray(serial_number))
    received = ser.read(19)
    if get_fcs(received, 4) != received[-2]: raise Exception("checksum is not equal")
    serial_num = int.from_bytes(received[13:17], "little")
    return serial_num

#reads article number of sensor
def get_article_number(adr : int) -> int:
    """Reads article number from sensor

    Args:
        adr (int): Sensor address

    Raises:
        Exception: Checksum is not equal

    Returns:
        int: Article number
    """
    article_number[4] = adr
    article_number[-2] = get_fcs(article_number, 4)
    ser.write(bytearray(article_number))
    received = ser.read(19)
    if get_fcs(received, 4) != received[-2]: raise Exception("checksum is not equal")
    article_n = int.from_bytes(received[13:17], "little")
    return article_n

#returns software version of sensor
def get_sw_version(adr : int) -> str:
    """Reads sw version from sensor

    Args:
        adr (int): Sensor address

    Raises:
        Exception: Checksum is not equal

    Returns:
        str: Sw version
    """
    sw_version[4] = adr
    sw_version[-2] = get_fcs(sw_version, 4)
    ser.write(bytearray(sw_version))
    received = ser.read(17)
    if get_fcs(received, 4) != received[-2]: raise Exception("checksum is not equal")
    sw = received[13:15]
    sw_string = f'0.{sw[1]}{chr(sw[0])}'
    return sw_string

#reads destriction of sensor
def get_description(adr : int) -> str:
    """Reads description from sensor

    Args:
        adr (int): Sensor address

    Raises:
        Exception: Checksum is not equal

    Returns:
        str: Description
    """
    description[4] = adr
    description[-2] = get_fcs(description, 4)
    ser.write(bytearray(description))
    received = ser.read(47)
    if get_fcs(received, 4) != received[-2]: raise Exception("checksum is not equal")
    temp = received[13:45]
    text = str(temp.decode("ascii"))
    text = text.strip()
    return text

#reads measuring unit of sensor
def get_measuring_unit(adr : int) -> str:
    """Reads measuring unit from sensor

    Args:
        adr (int): Sensor address

    Raises:
        Exception: Checksum is not equal

    Returns:
        str: Measuring unit
    """
    measuring_unit[4] = adr
    measuring_unit[-2] = get_fcs(measuring_unit, 4)
    ser.write(bytearray(measuring_unit))
    received = ser.read(16)
    if get_fcs(received, 4) != received[-2]: raise Exception("Checksum is not equal")
    unit = int(received[13])
    units = ["m","mm","μm"]
    return units[unit]

#reads measuring range of sensor
def get_measuring_range(adr : int) -> float:
    """Reads measuring range from sensor

    Args:
        adr (int): Sensor address

    Raises:
        Exception: Checksum is not equal

    Returns:
        float: Measuring range
    """
    measuring_range[4] = adr
    measuring_range[-2] = get_fcs(measuring_range, 4)
    ser.write(bytearray(measuring_range))
    received = ser.read(19)
    if get_fcs(received, 4) != received[-2]: raise Exception("checksum is not equal")
    range = struct.unpack('f', received[13:17])[0]
    return range

#reads measuring offset of sensor
def get_measuring_offset(adr : int) -> float:
    """Reads measuring offset from sensor

    Args:
        adr (int): Sensor address

    Raises:
        Exception: Checksum is not equal

    Returns:
        float: Measuring offset
    """
    measuring_offset[4] = adr
    measuring_offset[-2] = get_fcs(measuring_offset, 4)
    ser.write(bytearray(measuring_offset))
    received = ser.read(19)
    if get_fcs(received, 4) != received[-2]: raise Exception("checksum is not equal")
    offset = struct.unpack('f', received[13:17])[0]
    return offset

#returns address if it exists, or if adr=0x7f(broadcast), returns address of single connected sensor
def get_adr(adr : int) -> int:
    """Reads address from sensor. Returns address any single sensor if adr is broadcast(0x7f)

    Args:
        adr (int): Sensor address

    Raises:
        Exception: Checksum is not equal

    Returns:
        int: Address
    """
    broadcast[1] = adr
    broadcast[-2] = get_fcs(broadcast, 1)
    ser.write(bytearray(broadcast))
    received = ser.read(6)
    if get_fcs(received, 1) != received[-2]: raise Exception("Checksum is not equal")
    return received[2]

#retrieve all date from sensor
def get_all(adr: int) -> dict:
    """Reads all values from sensor

    Args:
        adr (int): Sensor address

    Returns:
        dict: Sensor readings
    """
    values = {}
    values["Address"] = adr
    values["Serial Number"] = get_serial_number(adr)
    values["SW Version"] = get_sw_version(adr)
    values["Article Number"] = get_article_number(adr) 
    values["Description"] = get_description(adr)
    values["Measuring Unit"] = get_measuring_unit(adr)
    values["Measuring Range"] = get_measuring_range(adr)
    values["Measuring Offset"] = get_measuring_offset(adr)
    values["Distance"] = get_distance(adr)
    values["Temperature"] = get_temperature(adr)  
    return values

def read_sensor(adr_func, sensor):
    try:
        txt_log.insert(END, "\n")
        try: log_text=f'Machine: {get_machine()},  Sensor: {sensor()}'
        except: log_text="No sensor selected"
        log_write(log_text)
        retrieved = get_all(adr_func())
        for key in retrieved:
            labels[key]['text'] = retrieved[key]
            log_write(f"{key}: {retrieved[key]}")
            try:
                sensor().values[key] = retrieved[key]
            except: pass
        try:
            treeview_dict[sensor().location][0] = retrieved["Serial Number"]
            treeview_dict[sensor().location][1] = retrieved["Address"]
            treeview_dict[sensor().location][3] = retrieved["Distance"]
            current_address_lbl_value['text'] = retrieved["Address"]
            update_treeview()
        except: pass
    except Exception as error:
        log_write(error)


#retrieves all data from sensors when there are mulitple sensors connected
def read_all_sensors():
    try:
        sensors_group = get_machine().sensors_group
        sensors_group.clear()
        ser.timeout=0.02
        for i in range(0, 127):
            try: 
                progress_var.set(i)
                window.update_idletasks()
                adr = get_adr(i)                
            except: pass
            else:
                txt_log.insert(END, "\n")
                log_write(f'found adr: {adr}')
                ser.timeout=1
                serial_num = get_serial_number(adr)
                sensor_found = False
                for sensor in get_machine().sensors:
                    if sensor.values["Serial Number"] == serial_num:
                        sensor_data = [sensor.location, sensor.address, sensor.nom_value, sensor.tolerance]                   
                        sensor_found = True
                        break                
                if not sensor_found: 
                    sensor_data = ["Unknown", "Unknown", "Unknown", "Unknown"]                                      
                    log_write(f"Sensor: {serial_num} not found in registry")  
                sensors_group.append(Sensor(sensor_data[0], sensor_data[1], sensor_data[2], sensor_data[3]))  
                read_sensor(lambda:adr, lambda:sensors_group[-1])                               
                ser.timeout=0.02  
        combobox_group_sensors['values'] = sensors_group
        current_group_sensor.set("")     
        combobox_group_sensors.current(newindex=len(sensors_group)-1)
        current_sensor.set("")
        ser.timeout=1
    except Exception as error: 
        log_write(error)
    progress_var.set(0)

#serial write functions
#changes the address of single connected sensor
def set_address(new_adr):
    try:
        try: current_adr = get_adr(get_sensor().values["Address"])
        except: raise Exception("Sensor is not present")
        assign_address[4] = current_adr
        assign_address[9] = new_adr
        assign_address[-2] = get_fcs(assign_address, 4)
        ser.write(bytearray(assign_address))
        received = ser.read(1)
        log_write(f'Received: {received}')
        messagebox.showinfo("Assign Address", "Restart the sensor then click 'ok'")
        seconds = 60
        wait = time.time() + seconds
        polling_sensor(new_adr, seconds, wait)
    except Exception as error:
        log_text=error
        log_write(log_text)

#looks for sensor after reset follwing address change
def polling_sensor(adr, seconds, wait):
    ser.timeout = 0.01
    if ser.in_waiting != 0: ser.reset_input_buffer()
    try:
        if(time.time() > wait):
            current_adr = "Sensor not found"                
        else: current_adr = get_adr(adr)              
    except: 
        time_left = wait - time.time()
        if(time_left < seconds):
            log_write(f'Looking for sensor: {seconds}')
            seconds -= 1
        window.after(100, lambda:polling_sensor(adr, seconds, wait))
    else:
        try: 
            get_sensor().values["Address"] = current_adr
            treeview_dict[get_sensor().location][1] = current_adr
        except: pass
        labels["Address"]['text'] = current_adr
        current_address_lbl_value['text'] = current_adr
        log_text=f'current address: {current_adr}'
        log_write(log_text)
        update_treeview()
    ser.timeout = 1

#Utility functions

#refreshes the values in list of ports
def refresh_ports(): 
    global ports
    global ser
    ports = serial.tools.list_ports.comports()
    combobox_ports['values'] = ports
    current_port.set("")
    ser = None

#writes string to log
def log_write(text : str): 
    current_time = datetime.datetime.now().strftime("%H:%M:%S")
    logging.info(f'{str(current_time)} {text}')
    txt_log.insert(END, f'{str(current_time)} {text}\n')    
    txt_log.yview_moveto(1.0)

#function is called when a port is selected from the list
def port_selected():
    global ser
    ser = None
    ser = initialize_port()
    window.focus()
    log_text=f'Selected port: {current_port.get()}'
    log_write(log_text)

#function is called when a machine is selected from the list
def machine_selected():
    global treeview_dict
    window.focus()
    log_text=f'Selected machine: {current_machine.get()}'
    log_write(log_text)
    combobox_sensors['values'] = get_machine().sensors
    combobox_group_sensors['values'] = get_machine().sensors_group
    current_sensor.set("")    
    for key in labels:
        labels[key]['text'] = ""
    if get_machine().machine_type == "Compressor":
        treeview_dict = get_machine().compressors
    elif get_machine().machine_type == "Pump":
        treeview_dict = get_machine().pumps
    update_treeview()

#function is called when a sensor is selected from the list
def sensor_selected():
    window.focus()
    current_group_sensor.set("")
    sensor = get_sensor()
    txt_log.insert(END, "\n")
    log_text=f'Selected sensor: {current_sensor.get()}'
    update_sensor_lbl()
    log_write(log_text) 
    #log_write(sensor)
    for key in sensor.values:
        if sensor.values[key]: log_write(f'{key}: {sensor.values[key]}')
        labels[key]['text'] = sensor.values[key]
    selected_address.set(sensor.address)
    current_address_lbl_value["text"] = sensor.values["Address"]

#function is called when a sensor is selected from the group sensor list
def group_sensor_selected():
    window.focus()
    current_sensor.set("")
    group_sensor = get_machine().sensors_group[combobox_group_sensors.current()]
    txt_log.insert(END, "\n")
    log_text=f'Selected sensor: {current_group_sensor.get()}'
    log_write(log_text) 
    #log_write(sensor)
    for key in group_sensor.values:
        if group_sensor.values[key]: log_write(f'{key}: {group_sensor.values[key]}')
        labels[key]['text'] = group_sensor.values[key]
    
#creates a gui button with label to display results
def create_button(text, func, padx_btn=(20,8), pady=(8,8)):
    frame = ttk.Frame(frame_right, width=50, height=50)
    btn = ttk.Button(frame, text=text, width=15, command=lambda: func())
    btn.pack(side=tk.LEFT, padx=padx_btn, pady=pady, ipady=1)
    frame.pack(anchor=W)

def create_display(text, top_frame, padx_lbl=(20,8), padx_lbl_value=(8,20), pady=(8,8)):
    frame = ttk.Frame(top_frame, width=50, height=50)
    lbl = ttk.Label(frame, text=text, width=14)
    lbl.pack(side=tk.LEFT, padx=padx_lbl, pady=pady)
    lbl_value = ttk.Label(frame, text="", width=14, background="white", relief=SOLID)
    lbl_value.pack(side=tk.RIGHT, padx=padx_lbl_value, pady=pady)
    frame.pack()
    return lbl_value

#calculates the fcs checksum of the telegram
def get_fcs(telegram, x):
    sum=0
    for i in range(x, len(telegram)-2):
        sum += telegram[i]
    return sum % 256

#initializes Serial object
def initialize_port():
    index = combobox_ports.current()
    if index < 0:
        raise Exception()
    #Creates Serial object
    COM = ports[index].device
    ser = serial.Serial(
        port=COM,
        baudrate=230400,
        bytesize=serial.EIGHTBITS,
        stopbits=serial.STOPBITS_ONE,
        parity=serial.PARITY_EVEN,
        timeout=1
    )
    return ser

#opens a dialog window for creating a new Machine object
def create_machine():
    dialog = New_Machine(parent=window, msg_list=["ITM Number:","Description:","Serial Number:"])
    window.wait_window(dialog.window)
    if dialog.cancel == False:
        machines.append(Machine(dialog.inputs[0],dialog.inputs[1],dialog.inputs[2],dialog.machine_type))
        combobox_machines['values'] = machines
        combobox_machines.current(newindex=len(machines)-1)
        log_write(f'Added machine with Itm Number: {machines[0].itm_number}')
        machine_selected()

#opens a dialog widow for creating a new Sensor object
def create_sensor():
    if combobox_machines.current() == -1: 
        messagebox.showinfo("error", "No selected machine")
        return
    dialog = New_Sensor(window, get_machine())
    window.wait_window(dialog.window)
    if dialog.cancel == False:
        sensors = machines[combobox_machines.current()].sensors
        values = dialog.values
        sensors.append(Sensor(values[0],values[1],values[2],values[3]))
        combobox_sensors['values'] = sensors
        combobox_sensors.current(newindex=len(sensors)-1)
        log_write(f'Added sensor with location: {values[0]}')
        sensor_selected()
        selected_address.set(values[1])

#returns the selected Machine object
def get_machine():
    index = combobox_machines.current()
    if index < 0: raise Exception("Machine not selected")
    return machines[index]

#returns the selected Sensor object
def get_sensor():
    sensor_index = combobox_sensors.current()
    if sensor_index < 0: raise Exception("Sensor not selected")
    return get_machine().sensors[sensor_index]

def update_sensor_lbl():
    try:
        lbl_current_sensor['text'] = f"Sensor: {get_sensor().values['Serial Number']} Location: {get_sensor().location}"
    except: pass

def reset_buffer():
    ser.reset_input_buffer()

def get_temp_corr():
    #reads temperature correction values from text file
    temp_corr = {}
    for line in open("TempCorrFactor.txt").readlines():
        if line != '\n':
            var = line.split()
            temp_corr[var[0]] = [float(var[1]), float(var[2])]
    return temp_corr

#saves machines list to pickle file
def save():
    Files = [('Pickle Files', '*.p')]
    file = filedialog.asksaveasfilename(filetypes = Files, defaultextension = Files)
    pickle.dump(machines, open(file, "wb"))
 
#loads machines list from pickle file
def load():
    global machines
    Files = [('Pickle Files', '*.p')]
    file = filedialog.askopenfilename(filetypes=Files, defaultextension = Files) 
    machines = pickle.load(open(file, "rb"))
    combobox_machines['values'] = machines

_loop = None
def loop():
    try:
        global _loop
        values = get_all(get_adr(0x7f))
        log_write(values)
    except Exception as error: 
        log_write(error) 
    _loop = window.after(50, loop)

def endloop():
    window.after_cancel(_loop)

def print_excel(machine):
    wb = load_workbook(filename = 'ME-ProxSensor-TestProgDataFile.xlsx')
    ws = wb.active
    print_machine(ws, machine)
    print_values(ws, machine.sensors, machine)
    print_values(ws, machine.sensors_group, machine)
    wb.save('test.xlsx')

def print_values(ws, sensors, machine):
    row = 19
    column = 1
    for sensor in sensors:
        while ws.cell(row, column).value:
            column += 1
        else:
            dataset = column
            if sensors is machine.sensors: index = 1
            else: index = 2
            print_dataset(ws, sensor, f'Dataset: {dataset}', index)
            ws.cell(row, column, str(column))
            row += 1
            for key in sensor.values:
                ws.cell(row, column, str(sensor.values[key]))
                row += 1
            row = 19
            column += 1

def print_dataset(ws, sensor, dataset, index):
    row = 6
    column = 1
    while ws.cell(row, column).value:
        if ws.cell(row, column).value == sensor.location:
            ws.cell(row, column + index, dataset)
            break
        row += 1

def print_machine(ws, machine):
    ws.cell(1, 1, "Project: test")
    ws.cell(2, 1, f'Machine ITM: {machine.itm_number}')
    ws.cell(3, 1, f'Machine Description {machine.description}')
    ws.cell(4, 1, f'Machine Serial No.: {machine.serial_number}')

def update_treeview():
    tree.delete(*tree.get_children())
    tag = 0
    for key in treeview_dict:
        items = [key] + treeview_dict[key][0:4]
        tree.insert('', tk.END, values=items, tag=str(tag))
        try:
            distance = float(tree.item(tree.get_children()[tag])['values'][4])
            default_address = int(tree.item(tree.get_children()[tag])['values'][3])
            current_address = int(tree.item(tree.get_children()[tag])['values'][2])
            nom_value = int(treeview_dict[key][4])
            tolerance = int(treeview_dict[key][5])
            if distance < nom_value + tolerance and distance > nom_value - tolerance:
                tree.tag_configure(str(tag), background="limegreen")  
                if current_address != default_address:
                    tree.tag_configure(str(tag), background="yellow")  
            else: tree.tag_configure(str(tag), background="red")        
        except: pass
        tag += 1

#Creates Window
ctypes.windll.shcore.SetProcessDpiAwareness(1)
window = tk.Tk()
window.title("OneSubsea ME Proximity sensor test")
window.geometry('1500x900+300+50')
window.resizable(False,False)
style = ttk.Style()
style.configure('Treeview', rowheight=25)

#Configures log file
logging.basicConfig(filename="Log.log", encoding='utf-8', level=logging.INFO, format='%(levelname)s:%(message)s')

#adds onesubsea icon and logo to window
canvas = tk.Canvas(window, width = 700, height = 150)
canvas.grid(row=0, column=1)
img = ImageTk.PhotoImage(Image.open("onesubsea_logo.png"))
canvas.create_image(20, 20, anchor=NW, image=img)
icon = ImageTk.PhotoImage(Image.open("onesubsea_icon.png"))
window.iconphoto(False, icon)

frame_center = ttk.Frame(master=window)
frame_right = ttk.Frame(master=window)
frame_left = ttk.Frame(master=window)

lbl_current_sensor = ttk.Label(frame_right, text="")
lbl_current_sensor.pack(padx=(8,8), pady=(0,5))

#Creates button for retreiving all data from sensor and a label to display result
create_button("Read Sensor", lambda:read_sensor(lambda:get_adr(0x7f), get_sensor), pady=(8,8))
#Creates button for retreiving temperature value from sensor and a label to display result
labels["Temperature"] = create_display("Temperature", frame_right)
#Creates button for retreiving distance value from sensor and a label to display result
labels["Distance"] = create_display("Distance",frame_right)
#Creates button for retreiving serial number from sensor and a label to display result
labels["Serial Number"] = create_display("Serial Number",frame_right)
#Creates button for retreiving software version from sensor and a label to display result
labels["SW Version"] = create_display("SW Version",frame_right)
#Creates button for retreiving article number from sensor and a label to display result
labels["Article Number"] = create_display("Article Number",frame_right)
#Creates button for retreiving description from sensor and a label to display result
labels["Description"] = create_display("Description",frame_right)
#Creates button for retreiving measuring unit from sensor and a label to display result
labels["Measuring Unit"] = create_display("Measuring Unit",frame_right)
#Creates button for retreiving measuring range from sensor and a label to display result
labels["Measuring Range"] = create_display("Measuring Range",frame_right)
#Creates button for retreiving measuring offset from sensor and a label to display result
labels["Measuring Offset"] = create_display("Measuring Offset",frame_right)
#Creates button for retreiving address from sensor and a label to display result
labels["Address"] = create_display("Address", frame_right)

#creates a button for retreiving all data from group of sensors
create_button("Read All Sensors", read_all_sensors)
create_button("Reset Buffer", reset_buffer)

#Creates a text box for displaying log
frame_log = ttk.Frame(frame_center)
txt_log = tk.Text(frame_log, height=15, width=61)
lbl_log = ttk.Label(frame_log, text="Log")
lbl_log.pack(anchor=NW)
txt_log.pack(side=LEFT)
scrollbar = ttk.Scrollbar(frame_log, orient=tk.VERTICAL, command=txt_log.yview)
txt_log.configure(yscroll=scrollbar.set)
scrollbar.pack(side=RIGHT, fill=tk.BOTH)

progress_var = tk.DoubleVar() #here you have ints but when calc. %'s usually floats
progressbar = ttk.Progressbar(frame_center, variable=progress_var, maximum=126, length=737)

#creates treeview
columns = ("one", "two", "three", "four", "five")
tree_frame = ttk.Frame(frame_center)
tree = ttk.Treeview(tree_frame, columns=columns, show='headings', selectmode='browse')
tree.heading('one', text="Location")
tree.heading('two', text="Serial Number")
tree.heading('three', text="Current Address")
tree.heading('four', text="Default Address")
tree.heading('five', text="Distance")
tree.column(column=columns[0], width=174)
tree.column(column=columns[1], anchor=CENTER, width=141)
for i in range(2,5): tree.column(column=columns[i], anchor=CENTER, width=140)
tree.pack(side=LEFT)
# add a scrollbar
scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
tree.configure(yscroll=scrollbar.set)
scrollbar.pack(side=RIGHT, fill=tk.BOTH)


address_frame = ttk.Frame(frame_left, width=50, height=50)
current_address_frame = ttk.Frame(address_frame, width=50, height=50)
current_address_frame.pack(anchor=NW)
current_address_lbl = ttk.Label(current_address_frame, text="Current Address", width=20)
current_address_lbl.pack(side=tk.TOP, padx=(20,0), pady=(50,0), anchor=NW)
current_address_lbl_value = ttk.Label(current_address_frame, text="", width=20, background="white", relief=SOLID)
current_address_lbl_value.pack(side=tk.BOTTOM, padx=(20,0), pady=(0,0), ipady=1)
#Creates a list and button for changing address
frame_new_address = ttk.Frame(address_frame, width=50, height=50)
frame_new_address.pack(anchor=SW)
lbl_new_address = ttk.Label(frame_new_address, text="New Address")
lbl_new_address.pack(anchor=NW, padx=(20,5), pady=(0,0))
selected_address = tk.StringVar()
entry_new_address = ttk.Entry(frame_new_address, textvariable=selected_address, width=20)
entry_new_address.pack(side=tk.LEFT, padx=(20,5), pady=(0,0))
change_address_btn = ttk.Button(frame_new_address, text="Change Address", command=lambda: set_address(int(selected_address.get())))
change_address_btn.pack(side=tk.RIGHT, padx=(5,20), pady=(0,0))


#Creates a label, dropdown list and refresh button for port selection
frame_ports = ttk.Frame(frame_left, width=50, height=50)
lbl_ports = ttk.Label(frame_ports, text="Available Ports")
lbl_ports.pack(anchor=NW, padx=(20,5), pady=(0,0))
current_port = tk.StringVar()
combobox_ports = ttk.Combobox(frame_ports, textvariable=current_port, width=18)
combobox_ports['state'] = 'readonly'
combobox_ports.bind("<<ComboboxSelected>>", lambda e: port_selected())
combobox_ports.pack(side=tk.LEFT, padx=(20,5), pady=(0,0))
ports = serial.tools.list_ports.comports() #Creates a list of connected serial ports
combobox_ports['values'] = ports #adds the "ports" list to a dropdown menu
if(len(ports) > 0): 
    combobox_ports.current(newindex=0)
    port_selected()
refresh_btn = ttk.Button(frame_ports, text="Refresh", command=refresh_ports)
refresh_btn.pack(side=tk.RIGHT, padx=(5,20), pady=(0,0))

#Creates a list for machine selection
frame_machines = ttk.Frame(frame_left, width=50, height=50)
lbl_machines = ttk.Label(frame_machines, text="Machines")
lbl_machines.pack(anchor=NW, padx=(20,5), pady=(50,0))
current_machine = tk.StringVar()
combobox_machines = ttk.Combobox(frame_machines, textvariable=current_machine, width=18)
combobox_machines['state'] = 'readonly'
combobox_machines.bind("<<ComboboxSelected>>", lambda e: machine_selected())
combobox_machines.pack(side=tk.LEFT, padx=(20,5), pady=(0,0))
add_machine_btn = ttk.Button(frame_machines, text="New Machine", command=create_machine)
add_machine_btn.pack(side=tk.RIGHT, padx=(5,20), pady=(0,0))

#Creates a list for sensor selection
frame_sensors = ttk.Frame(frame_left, width=50, height=50)
lbl_sensors = ttk.Label(frame_sensors, text="Sensors")
lbl_sensors.pack(anchor=NW, padx=(20,5), pady=(50,0))
current_sensor = tk.StringVar()
combobox_sensors = ttk.Combobox(frame_sensors, textvariable=current_sensor, width=18)
combobox_sensors['state'] = 'readonly'
combobox_sensors.bind("<<ComboboxSelected>>", lambda e: sensor_selected())
combobox_sensors.pack(side=tk.LEFT, padx=(20,5), pady=(0,0))
add_sensor_btn = ttk.Button(frame_sensors, text="New Sensor", command=create_sensor)
add_sensor_btn.pack(side=tk.RIGHT, padx=(5,20), pady=(0,0))

#Creates a list for group sensor selection
frame_group_sensors = ttk.Frame(frame_left, width=50, height=50)
lbl_group_sensors = ttk.Label(frame_group_sensors, text="Group Sensors")
lbl_group_sensors.pack(anchor=NW, padx=(20,5), pady=(50,0))
current_group_sensor = tk.StringVar()
combobox_group_sensors = ttk.Combobox(frame_group_sensors, textvariable=current_group_sensor, width=18)
combobox_group_sensors['state'] = 'readonly'
combobox_group_sensors.bind("<<ComboboxSelected>>", lambda e: group_sensor_selected())
combobox_group_sensors.pack(side=tk.LEFT, padx=(20,5), pady=(0,0))

 
frame = ttk.Frame(window)
frame.grid()

 
mainmenu = tk.Menu(frame)
mainmenu.add_command(label = "Save", command= save)  
mainmenu.add_command(label = "Load", command= load)
mainmenu.add_command(label = "Exit", command= window.destroy)
mainmenu.add_command(label = "Loop", command= loop)
mainmenu.add_command(label = "End Loop", command= endloop)
mainmenu.add_command(label = "Print", command=lambda: print_excel(get_machine()))

window.config(menu = mainmenu)

#Places widgets in window
frame_center.grid(row=1, column=1, sticky='n')
frame_right.grid(row=1, column=2, sticky='n')
frame_left.grid(row=1, column=0, sticky='n')

frame_ports.pack(anchor=W)
frame_machines.pack(anchor=W)
frame_sensors.pack(anchor=W)
frame_group_sensors.pack(anchor=W)
address_frame.pack(anchor=W)
frame_log.pack()
progressbar.pack(anchor=W)
tree_frame.pack()

window.mainloop()